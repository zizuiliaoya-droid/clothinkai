"""U06a 导入异步 Runner（run_import_batch）。

按 nfr-design-patterns.md P-U06a-01/02 实现：
- **NF-1**：per-row 事务内 ``SET LOCAL app.tenant_id``（事务级，绝不会话级 —— 防连接池串租）
- 双 session：bypass（元数据 / 失败 job / 汇总，系统级）+ app（per-row upsert，RLS 约束）
- **NF-4**：``worker_process_init`` 信号注册 Adapter（HTTP 进程注册 worker 看不到）
- **FB-E**：only_failed 用 import_job.raw_data 还原失败行，原地 UPDATE attempt_count
- **FB-C**：runner 持有 per-row 事务边界；adapter.upsert(session, tenant_id, actor_id) 不自 commit

成功 job 与业务记录同 per-row 事务（原子）；失败 job 用独立 bypass session 写（不被回滚带走）。
"""

from __future__ import annotations

import asyncio
import io
import logging
import time
from typing import Any
from uuid import UUID

import sentry_sdk
from celery import Task
from celery.signals import worker_process_init
from sqlalchemy import func, select, text, update

from app.core.celery_app import celery_app
from app.core.config import settings
from app.core.db import AsyncSessionApp, AsyncSessionBypass
from app.core.metrics import (
    import_batch_duration_seconds,
    import_batch_total,
    import_rows_total,
)
from app.core.tenancy import tenant_id_ctx
from app.modules.importer.exceptions import RowValidationError
from app.modules.importer.models import ImportBatch, ImportJob
from app.modules.importer.registry import ImportAdapterRegistry

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# NF-4：worker 进程注册 Adapter（HTTP 进程的注册 worker 看不到）
# ---------------------------------------------------------------------------


@worker_process_init.connect
def _register_adapters_in_worker(**_kwargs: Any) -> None:
    """Celery worker 子进程启动时注册所有 Adapter（NF-4）。

    与 main.py lifespan 的 register_import_adapters 调用同一函数，保证
    HTTP 进程与 worker 进程都能 ImportAdapterRegistry.get(source)。
    """
    try:
        from app.main import register_import_adapters

        register_import_adapters()
        log.info("import_adapters_registered_in_worker")
    except Exception as exc:  # noqa: BLE001
        log.exception("import_adapter_worker_registration_failed")
        sentry_sdk.capture_exception(exc)


# ---------------------------------------------------------------------------
# Celery 任务入口
# ---------------------------------------------------------------------------


@celery_app.task(
    bind=True,
    name="app.tasks.import_tasks.run_import_batch",
    queue="default",
)
def run_import_batch(
    self: Task, batch_id: str, only_failed: bool = False
) -> dict[str, Any]:
    """异步执行导入批次解析 + 行级 upsert。

    Args:
        batch_id: ImportBatch.id（字符串，Celery JSON 序列化）。
        only_failed: True = 仅重跑 import_job.failed 行（FB-E partial 重试）。
    """
    return asyncio.run(_run_with_engine_dispose(UUID(batch_id), only_failed))


async def _run_with_engine_dispose(
    batch_id: UUID, only_failed: bool
) -> dict[str, Any]:
    """包裹任务执行：结束时 dispose 异步引擎。

    Celery worker 每个任务用独立 ``asyncio.run()``（新事件循环）。异步引擎的连接池
    会缓存绑定到上一个（已关闭）事件循环的 asyncpg 连接，下个任务复用时会挂起/报错
    （表现为批次卡在 processing）。每个任务结束 dispose 引擎，保证下个任务拿到新循环
    的新连接。
    """
    from app.core.db import engine_app, engine_bypass

    try:
        return await _run_import_batch(batch_id, only_failed)
    finally:
        await engine_app.dispose()
        await engine_bypass.dispose()


# ---------------------------------------------------------------------------
# 主编排（双 session + per-row SET LOCAL）
# ---------------------------------------------------------------------------


async def _run_import_batch(batch_id: UUID, only_failed: bool) -> dict[str, Any]:
    # ── 1. 元数据读取 + 状态守卫（bypass，系统级）──
    async with AsyncSessionBypass() as meta_s:
        batch = await meta_s.get(ImportBatch, batch_id)
        if batch is None:
            return {"status": "not_found"}
        # runner 入口守卫：仅 processing 可执行（NF-3 防重复 / 已结束）
        if batch.status != "processing":
            log.warning(
                "import_batch_not_processing",
                extra={"batch_id": str(batch_id), "status": batch.status},
            )
            return {"status": "skipped_not_processing"}
        tenant_id = batch.tenant_id
        source = batch.source
        created_by = batch.created_by
        file_bucket = batch.file_bucket
        file_r2_key = batch.file_r2_key
        original_filename = batch.original_filename

    # ── 2. Adapter 取得（NF-4：worker 已注册；缺失 → batch.failed）──
    adapter = ImportAdapterRegistry.get(source)
    if adapter is None:
        await _mark_batch_failed(batch_id, "adapter_not_registered")
        import_batch_total.labels(source=source, status="failed").inc()
        return {"status": "failed", "reason": "adapter_not_registered"}

    mapping = await _load_mapping(source, tenant_id, batch.mapping_version)

    # ── 3. 取文件 + 解析（解析致命失败 → batch.failed，FB-E ①）──
    try:
        if only_failed:
            rows = await _load_failed_rows(batch_id)  # [(row_number, raw_data), ...]
        else:
            from app.core.attachment import attachment_service

            raw = attachment_service.get_object_bytes(file_bucket, file_r2_key)
            rows = _parse_rows(raw, original_filename)
    except Exception as exc:  # noqa: BLE001
        await _mark_batch_failed(batch_id, f"parse_error:{type(exc).__name__}")
        sentry_sdk.capture_exception(exc)
        import_batch_total.labels(source=source, status="failed").inc()
        return {"status": "failed", "reason": "parse_error"}

    # 行数上限三层防护 L3（NF-6：upload 时无法预知行数）
    if len(rows) > settings.IMPORT_MAX_ROWS:
        await _mark_batch_failed(batch_id, "too_many_rows")
        import_batch_total.labels(source=source, status="failed").inc()
        return {"status": "failed", "reason": "too_many_rows"}

    # ── 4. 逐行处理（tenant_id_ctx 供 audit；RLS 靠 per-row SET LOCAL）──
    tok = tenant_id_ctx.set(tenant_id)
    start = time.perf_counter()
    imported = failed = 0
    try:
        for row_number, row in rows:
            ok = await _process_one_row(
                adapter,
                row,
                row_number,
                mapping,
                batch_id=batch_id,
                tenant_id=tenant_id,
                actor_id=created_by,
            )
            if ok:
                imported += 1
                import_rows_total.labels(source=source, result="success").inc()
            else:
                failed += 1
                import_rows_total.labels(source=source, result="failed").inc()
    finally:
        tenant_id_ctx.reset(tok)
        import_batch_duration_seconds.labels(source=source).observe(
            time.perf_counter() - start
        )

    # ── 5. 汇总（bypass）──
    status = await _summarize_batch(batch_id, imported, failed, only_failed)
    import_batch_total.labels(source=source, status=status).inc()
    return {"status": status, "imported": imported, "failed": failed}


# ---------------------------------------------------------------------------
# 单行处理（NF-1 核心：per-row 事务内 SET LOCAL）
# ---------------------------------------------------------------------------


async def _process_one_row(
    adapter: Any,
    row: dict[str, Any],
    row_number: int,
    mapping: Any,
    *,
    batch_id: UUID,
    tenant_id: UUID,
    actor_id: UUID | None,
) -> bool:
    """每行独立事务 + per-row SET LOCAL（NF-1 防连接池串租）。

    成功 → 业务记录 + import_job(success) 同事务原子提交。
    失败 → 独立 bypass session 写 import_job(failed)（防被业务事务回滚带走）。

    import_job 写入用 ``ON CONFLICT(batch_id, row_number)``：首次跑插入（attempt_count=1），
    重试时原地更新（attempt_count+1）—— FB-E only_failed 与首跑统一逻辑。
    """
    try:
        parsed = adapter.parse_row(row, mapping)
        errs = adapter.validate(parsed)
        if errs:
            raise RowValidationError("; ".join(errs))

        async with AsyncSessionApp() as app_s:
            # NF-1：SET LOCAL（事务级），commit/rollback 后失效，绝不残留连接池。
            # 用 set_config(setting, value, is_local=true)：等价 SET LOCAL，但接受 bind
            # 参数（asyncpg 的 SET 语法不支持占位符 $1，会 syntax error）。
            await app_s.execute(
                text("SELECT set_config('app.tenant_id', :tid, true)"),
                {"tid": str(tenant_id)},
            )
            rid, _inserted = await adapter.upsert(
                parsed,
                session=app_s,
                tenant_id=tenant_id,
                actor_id=actor_id,
            )
            await _upsert_job(
                app_s,
                batch_id=batch_id,
                tenant_id=tenant_id,
                row_number=row_number,
                row=row,
                status="success",
                error_detail=None,
                target_resource_id=rid,
            )
            await app_s.commit()
        return True
    except Exception as exc:  # noqa: BLE001
        # 失败行用独立 bypass session 写（不被业务回滚带走，FB-C + U05 模式）
        await _write_job_failed_bypass(
            batch_id=batch_id,
            tenant_id=tenant_id,
            row_number=row_number,
            row=row,
            error_detail=_sanitize(exc),
        )
        return False


async def _upsert_job(
    session: Any,
    *,
    batch_id: UUID,
    tenant_id: UUID,
    row_number: int,
    row: dict[str, Any],
    status: str,
    error_detail: str | None,
    target_resource_id: UUID | None,
) -> None:
    """INSERT ... ON CONFLICT(batch_id,row_number) DO UPDATE（首跑插入 / 重试更新）。

    用 core ``pg_insert``（绕过 ORM 租户钩子，显式带 tenant_id）；RLS 由调用方
    SET LOCAL（app 会话）或 bypass 角色（失败写）保证。
    """
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    stmt = pg_insert(ImportJob.__table__).values(
        tenant_id=tenant_id,
        batch_id=batch_id,
        row_number=row_number,
        status=status,
        raw_data=row,
        error_detail=error_detail,
        target_resource_id=target_resource_id,
        attempt_count=1,
    )
    stmt = stmt.on_conflict_do_update(
        index_elements=["batch_id", "row_number"],
        set_={
            "status": stmt.excluded.status,
            "raw_data": stmt.excluded.raw_data,
            "error_detail": stmt.excluded.error_detail,
            "target_resource_id": stmt.excluded.target_resource_id,
            "attempt_count": ImportJob.attempt_count + 1,
            "updated_at": func.now(),
        },
    )
    await session.execute(stmt)


async def _write_job_failed_bypass(
    *,
    batch_id: UUID,
    tenant_id: UUID,
    row_number: int,
    row: dict[str, Any],
    error_detail: str,
) -> None:
    """失败行用独立 bypass session 写（不被业务事务回滚带走）。"""
    async with AsyncSessionBypass() as fail_s:
        await _upsert_job(
            fail_s,
            batch_id=batch_id,
            tenant_id=tenant_id,
            row_number=row_number,
            row=row,
            status="failed",
            error_detail=error_detail,
            target_resource_id=None,
        )
        await fail_s.commit()


# ---------------------------------------------------------------------------
# 文件解析（csv / openpyxl read_only）
# ---------------------------------------------------------------------------


def _parse_rows(raw: bytes, filename: str) -> list[tuple[int, dict[str, Any]]]:
    """解析 CSV / XLSX 为 [(row_number, {col: value}), ...]（row_number 从 1 起，不含表头）。

    - CSV：utf-8-sig（兼容 BOM）+ DictReader
    - XLSX：openpyxl ``read_only=True, data_only=True``（流式 + 读公式计算值，不执行宏）
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _parse_csv(raw)
    if name.endswith(".xlsx"):
        return _parse_xlsx(raw)
    raise ValueError(f"unsupported file extension: {filename}")


def _parse_csv(raw: bytes) -> list[tuple[int, dict[str, Any]]]:
    import csv

    # 平台导出常带前置空行（如生意参谋千牛）→ 跳过开头完全空白的行后再取表头
    lines = raw.decode("utf-8-sig").splitlines()
    start = 0
    while start < len(lines) and lines[start].replace(",", "").strip() == "":
        start += 1
    text_stream = io.StringIO("\n".join(lines[start:]))
    reader = csv.DictReader(text_stream)
    rows: list[tuple[int, dict[str, Any]]] = []
    for idx, record in enumerate(reader, start=1):
        # 去除 None 键（多余列）+ 统一字符串
        clean = {
            (k or "").strip(): ("" if v is None else str(v))
            for k, v in record.items()
            if k is not None
        }
        rows.append((idx, clean))
    return rows


def _parse_xlsx(raw: bytes) -> list[tuple[int, dict[str, Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(
        io.BytesIO(raw), read_only=True, data_only=True
    )
    try:
        ws = wb.active
        rows: list[tuple[int, dict[str, Any]]] = []
        header: list[str] = []
        row_number = 0
        for excel_row in ws.iter_rows(values_only=True):
            cells = [
                str(c).strip() if c is not None else "" for c in excel_row
            ]
            if not header:
                # 平台导出常带前置空行/标题行（如生意参谋千牛表头在第 5 行）：
                # 跳过完全空白的前置行，第一行非空行作为表头
                if all(c == "" for c in cells):
                    continue
                header = cells
                continue
            row_number += 1
            record = {
                (header[j] if j < len(header) and header[j] else f"col_{j}"): (
                    "" if cell is None else str(cell)
                )
                for j, cell in enumerate(excel_row)
            }
            rows.append((row_number, record))
        return rows
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# bypass session 辅助（元数据 / 失败行 / 汇总）
# ---------------------------------------------------------------------------


async def _load_mapping(source: str, tenant_id: UUID, version: int | None) -> Any:
    """读取 field_mapping（按 batch.mapping_version；无版本 → None 恒等映射）。"""
    if version is None:
        return None
    from app.modules.importer.models import FieldMapping

    async with AsyncSessionBypass() as s:
        stmt = select(FieldMapping).where(
            FieldMapping.tenant_id == tenant_id,
            FieldMapping.source == source,
            FieldMapping.version == version,
        )
        return (await s.execute(stmt)).scalar_one_or_none()


async def _load_failed_rows(batch_id: UUID) -> list[tuple[int, dict[str, Any]]]:
    """only_failed 重试：用 import_job.raw_data 还原失败行（FB-E ②）。"""
    async with AsyncSessionBypass() as s:
        stmt = (
            select(ImportJob.row_number, ImportJob.raw_data)
            .where(ImportJob.batch_id == batch_id, ImportJob.status == "failed")
            .order_by(ImportJob.row_number.asc())
        )
        result = (await s.execute(stmt)).all()
    return [(int(rn), dict(rd)) for rn, rd in result]


async def _mark_batch_failed(batch_id: UUID, reason: str) -> None:
    """解析致命失败 / adapter 缺失 / 超行数 → batch.failed（bypass）。"""
    async with AsyncSessionBypass() as s:
        await s.execute(
            update(ImportBatch)
            .where(ImportBatch.id == batch_id)
            .values(status="failed", error_summary=reason[:2000], updated_at=func.now())
            .execution_options(synchronize_session=False)
        )
        await s.commit()


async def _summarize_batch(
    batch_id: UUID, imported: int, failed: int, only_failed: bool
) -> str:
    """汇总段：更新 batch 计数 + 终态（completed / partial / failed）。

    - only_failed=False（首跑 / 整文件重试）：total_rows = imported+failed，直接覆盖计数
    - only_failed=True（partial 重试）：重算 import_job 当前成功/失败总数（原地更新后）
    """
    async with AsyncSessionBypass() as s:
        if only_failed:
            total_success = (
                await s.execute(
                    select(func.count())
                    .select_from(ImportJob)
                    .where(ImportJob.batch_id == batch_id, ImportJob.status == "success")
                )
            ).scalar_one()
            total_failed = (
                await s.execute(
                    select(func.count())
                    .select_from(ImportJob)
                    .where(ImportJob.batch_id == batch_id, ImportJob.status == "failed")
                )
            ).scalar_one()
            imported_n, failed_n = int(total_success), int(total_failed)
            total_rows = imported_n + failed_n
        else:
            imported_n, failed_n = imported, failed
            total_rows = imported + failed

        if failed_n == 0 and total_rows > 0:
            status = "completed"
        elif imported_n == 0:
            status = "failed"
        else:
            status = "partial"

        await s.execute(
            update(ImportBatch)
            .where(ImportBatch.id == batch_id)
            .values(
                status=status,
                total_rows=total_rows,
                imported=imported_n,
                failed=failed_n,
                error_summary=(f"{failed_n} 行失败" if failed_n else None),
                updated_at=func.now(),
            )
            .execution_options(synchronize_session=False)
        )
        await s.commit()
    return status


def _sanitize(exc: Exception) -> str:
    """脱敏行级错误信息（截断 + 仅类型 + message，不含 SQL / 栈）。"""
    msg = getattr(exc, "message", None) or str(exc)
    return f"{type(exc).__name__}: {msg}"[:1000]


__all__ = ["run_import_batch"]
