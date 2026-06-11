"""U17 集成测试：套装创建+拆分 + 用户偏好 + 报表导出 + RLS。"""

from __future__ import annotations

import io
from datetime import date
from typing import Any

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import ValidationError
from app.core.tenancy import tenant_id_ctx
from app.modules.product.bundle_schemas import BundleCreate, BundleItemIn
from app.modules.product.bundle_service import BundleService
from app.modules.report.bi_service import DEFAULT_BI_LAYOUT
from app.modules.report.export_service import ReportExportService
from app.modules.report.user_preference_service import UserPreferenceService

pytestmark = pytest.mark.asyncio


class TestBundle:
    async def test_create_and_split(
        self, session: AsyncSession, tenant_a: Any, factory: Any,
        follower_role: Any, product_factory: Any,
    ) -> None:
        tok = tenant_id_ctx.set(tenant_a.id)
        try:
            user = await factory.user(tenant_a, roles=[follower_role])
            style = await product_factory.style()
            sku_a = await product_factory.sku(style, sku_code="A1")
            sku_b = await product_factory.sku(style, sku_code="B1")
            svc = BundleService(session)
            bundle = await svc.create(BundleCreate(
                bundle_code="BD001", bundle_name="春季套装",
                items=[BundleItemIn(sku_id=sku_a.id, quantity=1),
                       BundleItemIn(sku_id=sku_b.id, quantity=2)],
            ), user)
            _, items = await svc.get_with_items(bundle.id)
            assert len(items) == 2
            # 卖 3 件 → A:3, B:6
            split = dict(await svc.split_quantities(bundle.id, 3))
            assert split[sku_a.id] == 3
            assert split[sku_b.id] == 6
        finally:
            tenant_id_ctx.reset(tok)

    async def test_cross_tenant_sku_rejected(
        self, session: AsyncSession, tenant_a: Any, tenant_b: Any, factory: Any,
        follower_role: Any, product_factory: Any,
    ) -> None:
        tok = tenant_id_ctx.set(tenant_a.id)
        try:
            user = await factory.user(tenant_a, roles=[follower_role])
            style_b = await product_factory.style(tenant=tenant_b)
            sku_b = await product_factory.sku(style_b, tenant=tenant_b, sku_code="XB")
            svc = BundleService(session)
            with pytest.raises(ValidationError):
                await svc.create(BundleCreate(
                    bundle_code="BD002", bundle_name="x",
                    items=[BundleItemIn(sku_id=sku_b.id, quantity=1)],
                ), user)
        finally:
            tenant_id_ctx.reset(tok)

    async def test_duplicate_sku_in_bundle_rejected(
        self, session: AsyncSession, tenant_a: Any, factory: Any,
        follower_role: Any, product_factory: Any,
    ) -> None:
        tok = tenant_id_ctx.set(tenant_a.id)
        try:
            user = await factory.user(tenant_a, roles=[follower_role])
            style = await product_factory.style()
            sku_a = await product_factory.sku(style, sku_code="DUP")
            svc = BundleService(session)
            with pytest.raises(ValidationError):
                await svc.create(BundleCreate(
                    bundle_code="BD003", bundle_name="x",
                    items=[BundleItemIn(sku_id=sku_a.id, quantity=1),
                           BundleItemIn(sku_id=sku_a.id, quantity=2)],
                ), user)
        finally:
            tenant_id_ctx.reset(tok)


class TestUserPreference:
    async def test_upsert_and_get_default(
        self, session: AsyncSession, tenant_a: Any, factory: Any, admin_role: Any,
    ) -> None:
        tok = tenant_id_ctx.set(tenant_a.id)
        try:
            user = await factory.user(tenant_a, roles=[admin_role])
            svc = UserPreferenceService(session)
            # 无 → 默认
            got = await svc.get_or_default(user.id, "bi_layout", DEFAULT_BI_LAYOUT)
            assert got == DEFAULT_BI_LAYOUT
            # upsert → 回显
            await svc.upsert(user, "bi_layout", {"cards": ["x"]})
            got2 = await svc.get_or_default(user.id, "bi_layout", DEFAULT_BI_LAYOUT)
            assert got2 == {"cards": ["x"]}
        finally:
            tenant_id_ctx.reset(tok)


class TestExport:
    async def test_export_production_xlsx_parseable(
        self, session: AsyncSession, tenant_a: Any,
    ) -> None:
        tok = tenant_id_ctx.set(tenant_a.id)
        try:
            resp = await ReportExportService(session).export(
                tenant_a.id, "production", (date(2026, 6, 1), date(2026, 6, 30))
            )
            body = b"".join([chunk async for chunk in resp.body_iterator])
            wb = load_workbook(io.BytesIO(body))
            ws = wb.active
            header = [c.value for c in next(ws.iter_rows())]
            assert "款号" in header and "净投产比" in header
        finally:
            tenant_id_ctx.reset(tok)

    async def test_export_invalid_type(
        self, session: AsyncSession, tenant_a: Any,
    ) -> None:
        from app.modules.report.exceptions import ReportExportTypeInvalidError

        tok = tenant_id_ctx.set(tenant_a.id)
        try:
            with pytest.raises(ReportExportTypeInvalidError):
                await ReportExportService(session).export(
                    tenant_a.id, "unknown", (date(2026, 6, 1), date(2026, 6, 30))
                )
        finally:
            tenant_id_ctx.reset(tok)
