"""BI/导出 API 契约及 bundle 公开路由下线测试。"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from types import SimpleNamespace
from uuid import uuid4

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook


@pytest.mark.api
@pytest.mark.asyncio
class TestBundleExportApiContract:
    async def test_bi_dashboard_requires_auth(self) -> None:
        from app.main import app

        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as ac:
            resp = await ac.get("/api/reports/bi")
        assert resp.status_code == 401

    async def test_export_requires_auth(self) -> None:
        from app.main import app

        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as ac:
            resp = await ac.get("/api/reports/production/export")
        assert resp.status_code == 401

    async def test_openapi_hides_bundle_and_exposes_bi_export(self) -> None:
        from app.main import app

        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as ac:
            resp = await ac.get("/api/openapi.json")
        assert resp.status_code == 200
        paths = resp.json().get("paths", {})
        assert not any(path.startswith("/api/bundles") for path in paths)
        assert "/api/reports/bi" in paths
        assert "/api/reports/bi/layout" in paths
        assert "/api/reports/{report_type}/export" in paths

    async def test_export_rows_are_complete_and_numeric(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        from app.modules.report import export_service as module

        dangerous_header = '=HYPERLINK("https://invalid.example","x")'
        dangerous_text = "  +SUM(1,1)"
        production_row = SimpleNamespace(
            style_code="S001", style_name=dangerous_text, pay_amount=Decimal("100"),
            refund_amount=Decimal("10"), return_rate=Decimal("0.1"),
            confirmed_amount=Decimal("90"), promo_cost=Decimal("20"),
            ad_spend=Decimal("30"), total_spend=Decimal("50"), add_cart_count=5,
            add_cart_cost=Decimal("10"), net_roi=Decimal("1.8"),
            unit_deal_cost=Decimal("12"), extra={dangerous_header: "3"},
        )
        store_rows = [
            SimpleNamespace(
                date=date(2026, 8, 3), visitors=10, pay_amount=Decimal("20"),
                pay_orders=2, ad_spend_total=Decimal("3"),
                zhitongche_spend=Decimal("1"), yinli_spend=None,
                extra={"成交人数": "2"},
            ),
            SimpleNamespace(
                date=date(2026, 8, 4), visitors=5, pay_amount=Decimal("8"),
                pay_orders=1, ad_spend_total=None,
                zhitongche_spend=Decimal("2"), yinli_spend=Decimal("4"),
                extra={"成交人数": "1"},
            ),
        ]
        work_row = SimpleNamespace(
            pr_name="PR甲", quote_count=10, in_schedule_count=1, urge_count=2,
            important_urge_count=3, overdue_count=4, publish_count=5,
            info_complete_count=4, info_complete_rate=Decimal("0.8"), cancel_count=1,
            recall_due_count=2, recall_success_count=1,
            recall_complete_rate=Decimal("0.5"), overdue_rate=Decimal("0.4"),
            month_complete_rate=Decimal("0.5"), hit_count=2,
            hit_rate=Decimal("0.4"), like_count=100, cost=Decimal("200"),
            cpl=Decimal("2"),
        )

        class FakeProductionService:
            def __init__(self, _session: object) -> None:
                pass

            async def get_report(self, *_args: object, **_kwargs: object) -> object:
                return SimpleNamespace(items=[production_row])

        class FakeStoreDailyService:
            def __init__(self, _session: object) -> None:
                pass

            async def get_dashboard(self, *_args: object) -> list[object]:
                return store_rows

        class FakeWorkProgressService:
            def __init__(self, _session: object) -> None:
                pass

            async def get_for_month(self, *_args: object) -> list[object]:
                return [work_row]

        monkeypatch.setattr(module, "ProductionService", FakeProductionService)
        monkeypatch.setattr(module, "StoreDailyService", FakeStoreDailyService)
        monkeypatch.setattr(module, "WorkProgressService", FakeWorkProgressService)
        service = module.ReportExportService(None)  # type: ignore[arg-type]
        period = (date(2026, 8, 3), date(2026, 8, 4))

        production_headers, production_rows = await service._fetch_rows(
            uuid4(), "production", period,
            exclude_brushing=True, season=None, granularity="day",
        )
        assert production_headers[:2] == ["货号", "款名"]
        assert production_headers[-1] == dangerous_header
        assert len(production_headers) == len(production_rows[0]) == 14
        assert isinstance(module._cell(production_rows[0][2]), float)
        assert isinstance(module._cell(production_rows[0][-1]), float)
        assert module._cell(dangerous_header) == f"'{dangerous_header}"
        assert module._cell(dangerous_text) == f"'{dangerous_text}"

        store_headers, grouped_store_rows = await service._fetch_rows(
            uuid4(), "store-daily", period,
            exclude_brushing=True, season=None, granularity="week",
        )
        assert store_headers[:7] == [
            "日期", "访客数", "支付额", "支付订单", "广告花费",
            "直通车花费", "引力魔方花费",
        ]
        assert store_headers[-1] == "成交人数"
        assert grouped_store_rows[0][1:4] == [15, Decimal("28"), 3]
        assert grouped_store_rows[0][-1] == Decimal("3")

        work_headers, work_rows = await service._fetch_rows(
            uuid4(), "work-progress", period,
            exclude_brushing=True, season=None, granularity="day",
        )
        assert len(work_headers) == len(work_rows[0]) == 20
        assert work_headers[1:9] == [
            "约篇量", "档期内", "催发", "重要催发", "超时", "发布量",
            "信息完整数", "信息完整率",
        ]
        assert work_rows[0][7:9] == [4, Decimal("0.8")]

        response = await service.export(
            uuid4(), "production", period, granularity="day"
        )
        body = b"".join([chunk async for chunk in response.body_iterator])
        worksheet = load_workbook(io.BytesIO(body), data_only=False).active
        assert worksheet.cell(1, 14).value == f"'{dangerous_header}"
        assert worksheet.cell(1, 14).data_type != "f"
        assert worksheet.cell(2, 2).value == f"'{dangerous_text}"
        assert worksheet.cell(2, 2).data_type != "f"
